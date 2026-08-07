"""Generate one-time staging passwords and a hash-only SQL update file."""

from __future__ import annotations

import argparse
import logging
import secrets
import string
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import text

from app.core.db import engine
from app.core.security import get_password_hash

PASSWORD_ALPHABET = string.ascii_letters + string.digits + "@#%+-_"
logger = logging.getLogger(__name__)


def temporary_password() -> str:
    while True:
        value = "".join(secrets.choice(PASSWORD_ALPHABET) for _ in range(18))
        if (
            any(char.islower() for char in value)
            and any(char.isupper() for char in value)
            and any(char.isdigit() for char in value)
            and any(not char.isalnum() for char in value)
        ):
            return value


def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def load_accounts() -> list[dict[str, str]]:
    statement = text(
        """
        SELECT
            u.id::text AS id,
            COALESCE(o.name, '平台') AS organization,
            COALESCE(u.full_name, '') AS full_name,
            u.email,
            u.role::text AS role,
            COALESCE(string_agg(DISTINCT cg.name, '、'), '') AS classes
        FROM "user" AS u
        LEFT JOIN organization AS o ON o.id = u.org_id
        LEFT JOIN student AS s ON s.user_id = u.id
        LEFT JOIN classgroup AS student_class ON student_class.id = s.class_id
        LEFT JOIN teacherclasslink AS tcl ON tcl.user_id = u.id
        LEFT JOIN classgroup AS teacher_class ON teacher_class.id = tcl.class_id
        LEFT JOIN LATERAL (
            SELECT student_class.name
            WHERE student_class.name IS NOT NULL
            UNION ALL
            SELECT teacher_class.name
            WHERE teacher_class.name IS NOT NULL
        ) AS cg ON true
        GROUP BY u.id, o.name, u.full_name, u.email, u.role
        ORDER BY COALESCE(o.name, ''), u.role::text, u.full_name, u.email
        """
    )
    with engine.connect() as connection:
        return [dict(row) for row in connection.execute(statement).mappings()]


def write_workbook(rows: list[dict[str, str]], path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "试运行账号"
    sheet.append(["学校", "班级", "角色", "姓名", "登录账号", "临时密码"])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="2E5BFF")
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append(
            [
                row["organization"],
                row["classes"],
                row["role"],
                row["full_name"],
                row["email"],
                row["password"],
            ]
        )
    for column, width in zip("ABCDEF", (20, 24, 22, 20, 36, 24), strict=True):
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_hash_updates(rows: list[dict[str, str]], path: Path) -> None:
    lines = ["BEGIN;"]
    for row in rows:
        lines.append(
            "UPDATE \"user\" SET hashed_password = "
            f"{sql_literal(row['hashed_password'])} WHERE id = "
            f"{sql_literal(row['id'])}::uuid;"
        )
    lines.extend(["COMMIT;", ""])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--sql-output", type=Path, required=True)
    args = parser.parse_args()
    if args.workbook.exists() or args.sql_output.exists():
        raise SystemExit("Refusing to overwrite an existing credential artifact")

    rows = load_accounts()
    for row in rows:
        row["password"] = temporary_password()
        row["hashed_password"] = get_password_hash(row["password"])
    write_workbook(rows, args.workbook)
    write_hash_updates(rows, args.sql_output)
    logger.info("Prepared %s account rotations", len(rows))
    logger.info("Workbook: %s", args.workbook)
    logger.info("Hash-only SQL: %s", args.sql_output)


if __name__ == "__main__":
    main()
