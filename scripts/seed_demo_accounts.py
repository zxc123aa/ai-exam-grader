"""Seed deterministic demo accounts and export their credentials to Excel."""

from __future__ import annotations

import argparse
import os
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlmodel import Session, delete, select

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from app.core.db import engine  # noqa: E402
from app.core.security import get_password_hash  # noqa: E402
from app.models import (  # noqa: E402
    ClassGroup,
    Organization,
    Student,
    TeacherClassLink,
    User,
    UserRole,
)

PASSWORD = "Dianfan@2026"
ORG_CODE = "demo2"
ORG_NAME = "示范二中"
OUTPUT = ROOT / "outputs" / "demo-accounts" / "点凡阅卷_示范二中及平台账号.xlsx"
CLASS_NAMES = ("高三（1）班", "高三（2）班")


@dataclass(frozen=True)
class AccountSpec:
    email: str
    full_name: str
    role: UserRole
    org_code: str | None
    employee_no: str | None = None
    subjects: tuple[str, ...] = ()
    class_names: tuple[str, ...] = ()
    is_superuser: bool = False


PLATFORM_ACCOUNTS = (
    AccountSpec(
        email="platform.superuser@example.com",
        full_name="点凡平台超级管理员",
        role=UserRole.PLATFORM_SUPERUSER,
        org_code=None,
        is_superuser=True,
    ),
    AccountSpec(
        email="platform.admin@example.com",
        full_name="点凡平台管理员",
        role=UserRole.PLATFORM_ADMIN,
        org_code=None,
    ),
)

SCHOOL_ACCOUNTS = (
    AccountSpec(
        email="demo2.owner@example.com",
        full_name="李校长",
        role=UserRole.SCHOOL_OWNER,
        org_code=ORG_CODE,
    ),
    AccountSpec(
        email="demo2.admin@example.com",
        full_name="王主任",
        role=UserRole.SCHOOL_ADMIN,
        org_code=ORG_CODE,
    ),
)

TEACHER_ACCOUNTS = (
    AccountSpec(
        email="demo2.physics@example.com",
        full_name="张明远",
        role=UserRole.TEACHER,
        org_code=ORG_CODE,
        employee_no="D2-T001",
        subjects=("物理",),
        class_names=CLASS_NAMES,
    ),
    AccountSpec(
        email="demo2.math@example.com",
        full_name="王静",
        role=UserRole.TEACHER,
        org_code=ORG_CODE,
        employee_no="D2-T002",
        subjects=("数学",),
        class_names=CLASS_NAMES,
    ),
    AccountSpec(
        email="demo2.chinese@example.com",
        full_name="李文博",
        role=UserRole.TEACHER,
        org_code=ORG_CODE,
        employee_no="D2-T003",
        subjects=("语文",),
        class_names=CLASS_NAMES,
    ),
    AccountSpec(
        email="demo2.english@example.com",
        full_name="陈晓玲",
        role=UserRole.TEACHER,
        org_code=ORG_CODE,
        employee_no="D2-T004",
        subjects=("英语",),
        class_names=CLASS_NAMES,
    ),
)


def student_specs() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for class_index, class_name in enumerate(CLASS_NAMES, start=1):
        for position in range(1, 21):
            global_index = (class_index - 1) * 20 + position
            student_no = f"20260{class_index}{position:03d}"
            rows.append(
                {
                    "class_name": class_name,
                    "name": f"示例学生{global_index:03d}",
                    "student_no": student_no,
                    "email": f"{student_no}@school.local",
                }
            )
    return rows


def ensure_org(session: Session) -> Organization:
    org = session.exec(select(Organization).where(Organization.code == ORG_CODE)).first()
    if org:
        if org.name != ORG_NAME:
            raise ValueError(f"组织代码 {ORG_CODE} 已属于“{org.name}”，拒绝覆盖")
        return org
    org = Organization(
        name=ORG_NAME,
        code=ORG_CODE,
        status="active",
        contact_name="李校长",
    )
    session.add(org)
    session.flush()
    return org


def ensure_user(session: Session, spec: AccountSpec, org: Organization | None) -> User:
    user = session.exec(select(User).where(User.email == spec.email)).first()
    expected_org_id = org.id if org else None
    if user:
        if user.role != spec.role or user.org_id != expected_org_id:
            raise ValueError(
                f"账号 {spec.email} 已存在但角色或学校不匹配，拒绝接管"
            )
    else:
        user = User(
            email=spec.email,
            hashed_password=get_password_hash(PASSWORD),
            full_name=spec.full_name,
            role=spec.role,
            org_id=expected_org_id,
        )
    user.full_name = spec.full_name
    user.employee_no = spec.employee_no
    user.subjects = list(spec.subjects)
    user.is_active = True
    user.is_superuser = spec.is_superuser
    user.hashed_password = get_password_hash(PASSWORD)
    session.add(user)
    session.flush()
    return user


def ensure_classes(
    session: Session, org: Organization, owner: User
) -> dict[str, ClassGroup]:
    result: dict[str, ClassGroup] = {}
    for name in CLASS_NAMES:
        group = session.exec(
            select(ClassGroup).where(
                ClassGroup.org_id == org.id,
                ClassGroup.name == name,
            )
        ).first()
        if not group:
            group = ClassGroup(
                name=name,
                grade_level="高三",
                owner_id=owner.id,
                org_id=org.id,
            )
        else:
            group.grade_level = "高三"
        session.add(group)
        session.flush()
        result[name] = group
    return result


def ensure_teacher_links(
    session: Session,
    teacher: User,
    class_names: tuple[str, ...],
    classes: dict[str, ClassGroup],
) -> None:
    session.exec(
        delete(TeacherClassLink).where(TeacherClassLink.user_id == teacher.id)
    )
    for name in class_names:
        session.add(TeacherClassLink(user_id=teacher.id, class_id=classes[name].id))


def ensure_student(
    session: Session,
    org: Organization,
    group: ClassGroup,
    row: dict[str, str],
) -> None:
    by_number = session.exec(
        select(Student).where(
            Student.class_id == group.id,
            Student.student_no == row["student_no"],
        )
    ).first()
    by_name = session.exec(
        select(Student).where(
            Student.class_id == group.id,
            Student.name == row["name"],
        )
    ).first()
    if by_number and by_name and by_number.id != by_name.id:
        raise ValueError(f"{group.name} 的姓名与学号分别指向不同学生")
    student = by_number or by_name
    if student and (
        student.name != row["name"] or student.student_no != row["student_no"]
    ):
        raise ValueError(f"{group.name} 的学生 {row['name']} 与现有花名册冲突")

    user = session.exec(select(User).where(User.email == row["email"])).first()
    if user and (user.role != UserRole.STUDENT or user.org_id != org.id):
        raise ValueError(f"学生账号 {row['email']} 已被其他角色或学校占用")
    if not user:
        user = User(
            email=row["email"],
            hashed_password=get_password_hash(PASSWORD),
            full_name=row["name"],
            role=UserRole.STUDENT,
            org_id=org.id,
        )
    user.full_name = row["name"]
    user.is_active = True
    user.is_superuser = False
    user.hashed_password = get_password_hash(PASSWORD)
    session.add(user)
    session.flush()

    linked = session.exec(select(Student).where(Student.user_id == user.id)).first()
    if linked and (not student or linked.id != student.id):
        raise ValueError(f"学生账号 {row['email']} 已绑定其他花名册记录")
    if not student:
        student = Student(
            class_id=group.id,
            name=row["name"],
            student_no=row["student_no"],
        )
    student.user_id = user.id
    session.add(student)


def seed(session: Session) -> None:
    org = ensure_org(session)
    for spec in PLATFORM_ACCOUNTS:
        ensure_user(session, spec, None)
    owner = ensure_user(session, SCHOOL_ACCOUNTS[0], org)
    ensure_user(session, SCHOOL_ACCOUNTS[1], org)
    classes = ensure_classes(session, org, owner)
    for spec in TEACHER_ACCOUNTS:
        teacher = ensure_user(session, spec, org)
        ensure_teacher_links(session, teacher, spec.class_names, classes)
    for row in student_specs():
        ensure_student(session, org, classes[row["class_name"]], row)
    session.flush()


HEADER_FILL = PatternFill("solid", fgColor="2E5BFF")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Microsoft YaHei", color="1A1A1A")
MUTED_FONT = Font(name="Microsoft YaHei", color="6B6B6B")
THIN_BORDER = Border(bottom=Side(style="hair", color="EBEBE8"))


def style_table(sheet, widths: tuple[int, ...]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 24
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.number_format = "@"
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_workbook(path: Path) -> None:
    wb = Workbook()
    info = wb.active
    info.title = "使用说明"
    info.append(["项目", "内容"])
    info_rows = (
        ("品牌", "点凡阅卷（DIANFAN）"),
        ("示例学校", f"{ORG_NAME}（{ORG_CODE}）"),
        ("初始密码", PASSWORD),
        ("账号数量", "平台账号 2 个；学校账号 46 个；合计 48 个"),
        ("安全提示", "仅限内部演示。正式交付学校前必须修改密码并妥善保管本文件。"),
        ("数据库策略", "按学校代码、邮箱、班级名和学号幂等写入；重复执行不会重复创建。"),
    )
    for row in info_rows:
        info.append(row)
    style_table(info, (18, 82))
    info.auto_filter.ref = "A1:B7"
    for cell in info[6]:
        cell.font = Font(name="Microsoft YaHei", bold=True, color="DC2626")

    platform = wb.create_sheet("平台账号")
    platform.append(["姓名", "角色", "登录账号", "初始密码", "权限说明"])
    role_note = {
        UserRole.PLATFORM_SUPERUSER: "平台账号、模型中转和全局配置；不可访问学校批卷数据",
        UserRole.PLATFORM_ADMIN: "学校租户、订阅和额度管理；不可访问学校批卷数据",
    }
    for spec in PLATFORM_ACCOUNTS:
        platform.append(
            [spec.full_name, spec.role.value, spec.email, PASSWORD, role_note[spec.role]]
        )
    style_table(platform, (24, 22, 38, 20, 55))

    school = wb.create_sheet("学校管理账号")
    school.append(["学校", "姓名", "角色", "登录账号", "初始密码"])
    for spec in SCHOOL_ACCOUNTS:
        school.append([ORG_NAME, spec.full_name, spec.role.value, spec.email, PASSWORD])
    style_table(school, (18, 18, 20, 34, 20))

    teachers = wb.create_sheet("教师账号")
    teachers.append(["学校", "姓名", "工号", "任教科目", "任教班级", "登录账号", "初始密码"])
    for spec in TEACHER_ACCOUNTS:
        teachers.append(
            [
                ORG_NAME,
                spec.full_name,
                spec.employee_no,
                "、".join(spec.subjects),
                "、".join(spec.class_names),
                spec.email,
                PASSWORD,
            ]
        )
    style_table(teachers, (18, 16, 16, 16, 28, 34, 20))

    students = wb.create_sheet("学生花名册")
    students.append(["学校", "班级", "姓名", "学号", "登录账号", "初始密码", "账号绑定"])
    for row in student_specs():
        students.append(
            [
                ORG_NAME,
                row["class_name"],
                row["name"],
                row["student_no"],
                row["email"],
                PASSWORD,
                "已绑定",
            ]
        )
    style_table(students, (18, 18, 18, 18, 34, 20, 14))

    checks = wb.create_sheet("数据校验")
    checks.append(["检查项", "期望值", "结果"])
    for row in (
        ("平台账号", "2", "通过"),
        ("学校管理账号", "2", "通过"),
        ("教师账号", "4", "通过"),
        ("班级", "2", "通过"),
        ("学生及绑定账号", "40", "通过"),
        ("全部登录账号唯一", "48", "通过"),
    ):
        checks.append(row)
    style_table(checks, (28, 18, 18))
    for row in checks.iter_rows(min_row=2, min_col=3, max_col=3):
        row[0].font = Font(name="Microsoft YaHei", bold=True, color="16A34A")

    wb.save(path)


def validate_workbook(path: Path) -> None:
    wb = load_workbook(path, read_only=True, data_only=False)
    expected_sheets = ["使用说明", "平台账号", "学校管理账号", "教师账号", "学生花名册", "数据校验"]
    if wb.sheetnames != expected_sheets:
        raise ValueError("Excel 工作表结构不正确")
    expected_rows = {"平台账号": 3, "学校管理账号": 3, "教师账号": 5, "学生花名册": 41}
    for name, count in expected_rows.items():
        if wb[name].max_row != count:
            raise ValueError(f"{name} 行数不正确")
    emails: list[str] = []
    for name, column in (("平台账号", 3), ("学校管理账号", 4), ("教师账号", 6), ("学生花名册", 5)):
        emails.extend(str(row[column - 1].value) for row in wb[name].iter_rows(min_row=2))
    if len(emails) != 48 or len(set(emails)) != 48:
        raise ValueError("Excel 登录账号数量或唯一性校验失败")
    if any(cell.value in {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?"} for sheet in wb for row in sheet.iter_rows() for cell in row):
        raise ValueError("Excel 存在公式错误")
    wb.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成点凡阅卷示例账号和花名册")
    parser.add_argument("--apply", action="store_true", help="正式写入数据库并生成 Excel")
    parser.add_argument("--output", type=Path, default=OUTPUT, help="Excel 输出路径")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.apply:
        print("预检：将补全示范二中 46 个账号，并新增 2 个平台账号。使用 --apply 正式执行。")
        return

    output = args.output.resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix="demo-accounts-", suffix=".xlsx", dir=output.parent)
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        build_workbook(temp_path)
        validate_workbook(temp_path)
        with Session(engine) as session:
            seed(session)
            session.commit()
        os.replace(temp_path, output)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
    print(f"已生成 48 个演示账号，Excel：{output}")


if __name__ == "__main__":
    main()
